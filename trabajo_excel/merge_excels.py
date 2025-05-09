import openpyxl

def merge_excels(base_file, add_file, output_file, insert_position=0):
    """
    Combina las hojas de un archivo de Excel en otro, insertándolas en una posición específica.
    
    :param base_file: Ruta del archivo Excel base.
    :param add_file: Ruta del archivo Excel que se añadirá.
    :param output_file: Ruta donde se guardará el archivo combinado.
    :param insert_position: Índice de posición donde insertar las hojas del archivo adicional (0 es al inicio).
    """
    # Cargar los archivos Excel
    base_wb = openpyxl.load_workbook(base_file)
    add_wb = openpyxl.load_workbook(add_file)
    
    # Obtener las hojas del archivo base
    base_sheets = base_wb.sheetnames
    
    # Insertar hojas del archivo adicional
    for sheet_name in add_wb.sheetnames:
        add_sheet = add_wb[sheet_name]
        new_sheet = base_wb.create_sheet(title=sheet_name, index=insert_position)
        for row in add_sheet.iter_rows():
            for cell in row:
                new_sheet[cell.coordinate].value = cell.value
        insert_position += 1  # Incrementar la posición para mantener el orden

    # Guardar el archivo combinado
    base_wb.save(output_file)
    print(f"Archivo combinado guardado en: {output_file}")

# Ejemplo de uso
merge_excels(
    base_file="resultados.xlsx",
    add_file="resultados_bootstrap_fio2.xlsx",
    output_file="resultado.xlsx",
 )
